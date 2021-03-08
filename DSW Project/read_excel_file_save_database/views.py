from django.shortcuts import render
from django.http import HttpResponse
import re
# Create your views here.

import openpyxl
from read_excel_file_save_database.models import student

def index(request):
    if "GET" == request.method:
        return render(request, 'index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        for worksheets in wb.sheetnames:
            worksheet = wb[worksheets]
            print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        row_no=0
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                if cell.value!=None :
                    row_data.append(str(cell.value))
            if not row_data or row_no==0:
                excel_data.append(row_data)
                row_no+=1
                continue
            st = student()
            st.reg = row_data[0]
            st.ses = row_data[1]
            st.sem= row_data[2]
            st.sem_type = row_data[3]
            st.prog = row_data[4]
            st.bran = row_data[5]
            st.spi = row_data[6]
            st.p_cpi = row_data[7]
            st.cpi = row_data[8]
            st.result = row_data[9]
            n = 4
            myset = {"DN","BM","BT","CL","EN","GE","ST","CM","CC","CS","EE","PE","EM","FE","GI","IS","MT","PR","TR","VL","PS","PD","TH","SP","SW"}
            split = [st.reg[i:i+n] for i in range(0, len(st.reg), n)]
            match = re.match(r'.*([1-3][0-9]{3})', split[0])
            if match is not None :
                if split[1] == "PTSW" :
                    st.save()
                    excel_data.append(row_data)
                else :
                    n1=2
                    str1=split[1]
                    split1 = [str1[i:i+n1] for i in range(0, len(str1), n1)]
                    if split1[0] in myset :
                        st.save()
                        excel_data.append(row_data)
                    else :
                        row_data1 = list()
                        row_data1.append(str("Invalid Data of row no. "+ str(row_no)))
                        excel_data.append(row_data1)
            else :
                row_data1 = list()
                row_data1.append(str("Invalid Data of row no. "+ str(row_no)))
                excel_data.append(row_data1)
            row_no+=1

        return render(request, 'index.html', {"excel_data":excel_data})
